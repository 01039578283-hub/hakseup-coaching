from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook
from PIL import Image

try:
    from source_copy_utils import (
        VERIFIED_SCHOOL_SOURCE_CORRECTIONS,
        distribute_source_paragraphs,
        normalize_location_note,
        source_paragraphs,
        source_theme,
    )
except ModuleNotFoundError:  # package import
    from .source_copy_utils import (
        VERIFIED_SCHOOL_SOURCE_CORRECTIONS,
        distribute_source_paragraphs,
        normalize_location_note,
        source_paragraphs,
        source_theme,
    )


ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path.home() / "Desktop"
COMMON = DESKTOP / "홈페이지 정리" / "참고자료" / "공통자료"
DEFAULT_WORKBOOK = DESKTOP / "구글시트로 뽑은거" / "영수학원 원고.xlsx"
CENTER_CSV = COMMON / "센터정보 정리.csv"
IMAGE_CSV = COMMON / "이미지링크.csv"
NATIONAL_ROOT = ROOT / "전국학원"
SUBJECT_ROOT = ROOT / "과목별학원"
TARGET_ROOT = SUBJECT_ROOT / "영수학원"

BASE_URL = "https://xn--ru4bi8s1tac0p.kr"
SITE_NAME = "학습코칭 학원 안내"
CATEGORY = "영수학원"
PHONE = "010-3957-8283"
SMS_URL = "https://blogsms.net/01039578283"
FORM_URL = (
    "https://docs.google.com/forms/d/e/"
    "1FAIpQLSdb2oE5Qk5YS0TfYDxyV1w-IOTkhkjOCmmpAKTI9FmqpVj6Yg/viewform"
)
TODAY = date.today().isoformat()
PUBLISHED_DATE = "2026-08-13"

GRADE_ORDER = [
    "초1", "초2", "초3", "초4", "초5", "초6",
    "중1", "중2", "중3", "고1", "고2", "고3",
]
GRADE_EXPANDED = {
    **{f"초{i}": f"초등학교 {i}학년" for i in range(1, 7)},
    **{f"중{i}": f"중학교 {i}학년" for i in range(1, 4)},
    **{f"고{i}": f"고등학교 {i}학년" for i in range(1, 4)},
}
REGION_ORDER = [
    "서울", "경기", "인천", "부산", "대구", "광주", "대전",
    "울산", "세종", "강원", "충청", "전라", "경상", "제주",
]


def esc(value: object) -> str:
    return html.escape(str(value or ""), quote=True)


def compact(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def stable_unique(values: list[str]) -> list[str]:
    return list(dict.fromkeys(value for value in values if value))


def stable_number(key: str, salt: str = "") -> int:
    digest = hashlib.sha256(f"{key}|{salt}".encode("utf-8")).hexdigest()
    return int(digest[:16], 16)


def choose(key: str, values: list[str] | tuple[str, ...], salt: str = "") -> str:
    return values[stable_number(key, salt) % len(values)]


def absolute_url(*parts: str) -> str:
    route = "/" + "/".join(part.strip("/") for part in parts if part) + "/"
    return BASE_URL + quote(route, safe="/")


def normalize_slug(value: object) -> str:
    return re.sub(r"[\s-]+", "", compact(value))


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def field(row: dict[str, str], starts_with: str, contains: str = "") -> str:
    for key, value in row.items():
        normalized = compact(key)
        if normalized.startswith(starts_with) and (not contains or contains in normalized):
            return compact(value)
    return ""


def grade_values(value: object) -> list[str]:
    found = re.findall(r"(?:초[1-6]|중[1-3]|고[1-3])", compact(value))
    return [grade for grade in GRADE_ORDER if grade in found]


def with_josa(value: str, consonant: str, vowel: str) -> str:
    text = compact(value)
    if not text:
        return text
    code = ord(text[-1])
    has_batchim = 0xAC00 <= code <= 0xD7A3 and (code - 0xAC00) % 28 != 0
    return text + (consonant if has_batchim else vowel)


def school_values(*values: object) -> list[str]:
    result: list[str] = []
    for raw in values:
        raw_text = str(raw or "").strip()
        text = compact(raw_text)
        if not text:
            continue
        if "모든 고등학교" in text or "상담 확인 필요" in text:
            continue
        verified = VERIFIED_SCHOOL_SOURCE_CORRECTIONS.get(text)
        if verified:
            result.extend(verified)
            continue
        # Outside the exact, reviewed manifest, split only punctuation and
        # line-break delimiters that are present in the source.  Whitespace
        # and school-name suffixes are never sufficient evidence of a boundary.
        for token in re.split(r"[,，./|·;\r\n]+", raw_text):
            token = compact(token)
            if not token or "모든 고등학교" in token or "상담 확인 필요" in token:
                continue
            if not re.search(r"(?:초등학교|중학교|고등학교|초|중|고)$", token):
                continue
            result.append(token)
    return stable_unique(result)


def physical_address_parts(address: str) -> tuple[str, str]:
    normalized = compact(address)
    first = normalized.split(" ", 1)[0] if normalized else ""
    region_aliases = {
        "서울특별시": "서울", "서울": "서울",
        "부산광역시": "부산", "부산": "부산",
        "대구광역시": "대구", "대구": "대구",
        "인천광역시": "인천", "인천": "인천",
        "광주광역시": "광주", "광주": "광주",
        "대전광역시": "대전", "대전": "대전",
        "울산광역시": "울산", "울산": "울산",
        "세종특별자치시": "세종특별자치시",
        "경기도": "경기도", "강원특별자치도": "강원특별자치도",
        "충청북도": "충청북도", "충청남도": "충청남도",
        "전북특별자치도": "전북특별자치도", "전라북도": "전라북도",
        "전라남도": "전라남도", "경상북도": "경상북도",
        "경상남도": "경상남도", "제주특별자치도": "제주특별자치도",
    }
    region = region_aliases.get(first, first)
    if first == "세종특별자치시":
        return region, "세종시"
    tokens = normalized.split()[1:5]
    city = next((token for token in tokens if re.search(r"(?:시|군|구)$", token)), "")
    return region, city


def service_area(row: dict[str, str], address: str) -> tuple[str, str]:
    region = compact(row.get("지역"))
    city = compact(row.get("시or구"))
    physical_region, physical_city = physical_address_parts(address)
    if city.endswith(("로", "길")) or "세종특별자치시" in address:
        return "세종", "세종시"
    return region or physical_region, city or physical_city


def image_size(path: Path) -> tuple[int, int]:
    with Image.open(path) as image:
        return image.size


def extract_json_graph(source: str) -> list[dict]:
    match = re.search(
        r'<script\b[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        source,
        re.I | re.S,
    )
    if not match:
        return []
    data = json.loads(match.group(1))
    return data.get("@graph", []) if isinstance(data, dict) else []


def has_schema_type(node: dict, expected: str) -> bool:
    value = node.get("@type", [])
    return expected in value if isinstance(value, list) else value == expected


def load_national_sources() -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for index_path in NATIONAL_ROOT.rglob("index.html"):
        folder = index_path.parent
        if not (folder / "초등영수학원" / "index.html").exists():
            continue
        relative = folder.relative_to(NATIONAL_ROOT).parts
        if len(relative) != 3:
            continue
        source = index_path.read_text(encoding="utf-8")
        canonical = re.search(r'<link\s+rel=["\']canonical["\']\s+href=["\']([^"\']+)', source, re.I)
        representative = re.search(r'<meta\s+property=["\']og:image["\']\s+content=["\']([^"\']+)', source, re.I)
        map_match = re.search(r'assets/maps/([^"\']+)', source, re.I)
        graph = extract_json_graph(source)
        org = next(
            (node for node in graph if isinstance(node, dict) and has_schema_type(node, "EducationalOrganization")),
            {},
        )
        key = normalize_slug(relative[-1])
        result[key] = {
            "source_url": canonical.group(1) if canonical else absolute_url("전국학원", *relative),
            "representative": representative.group(1) if representative else "",
            "map_file": map_match.group(1) if map_match else "",
            "organization_id": compact(org.get("@id")),
        }
    return result


def workbook_cells(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    cells = [str(row[0]) for row in sheet.iter_rows(values_only=True) if row and row[0]]
    workbook.close()
    return cells


@dataclass(frozen=True)
class Record:
    index: int
    locality: str
    slug: str
    service_region: str
    service_city: str
    physical_region: str
    physical_city: str
    center_name: str
    legal_name: str
    registration: str
    address: str
    location_note: str
    tuition_url: str
    english_grades: tuple[str, ...]
    math_grades: tuple[str, ...]
    common_grades: tuple[str, ...]
    schools: tuple[str, ...]
    source_text: str
    source_html: str
    source_url: str
    organization_id: str
    representative: str
    body_file: str
    body_mobile_file: str
    map_file: str
    body_width: int
    body_height: int
    map_width: int
    map_height: int
    english_focus: str
    english_evidence: str
    math_focus: str
    math_evidence: str
    selected_grade: str
    persona: str

    @property
    def key(self) -> str:
        return f"{self.index}|{self.locality}|{hashlib.sha256(self.source_text.encode()).hexdigest()}"

    @property
    def title(self) -> str:
        return f"{self.locality} 영수학원"

    @property
    def page_url(self) -> str:
        return absolute_url("과목별학원", CATEGORY, self.slug)


def detect_english_focus(text: str, key: str) -> tuple[str, str]:
    candidates = [
        ("듣기", "듣기 핵심어와 문장 이해", "놓친 구간을 표시하고 다시 들은 뒤 한 문장으로 요약한 기록"),
        ("청취", "듣기 핵심어와 문장 이해", "놓친 구간을 표시하고 다시 들은 뒤 한 문장으로 요약한 기록"),
        ("문법", "문법 개념의 문장 적용", "개념을 고른 이유와 실제 문장에서 바르게 고친 흔적"),
        ("서술형", "서술형 답안의 근거와 표현", "답의 근거를 표시하고 문장을 다시 완성한 기록"),
        ("어휘", "어휘 누적과 문장 적용", "단어 뜻만 외우지 않고 예문에서 다시 사용한 기록"),
        ("단어", "어휘 누적과 문장 적용", "단어 뜻만 외우지 않고 예문에서 다시 사용한 기록"),
        ("독해", "독해 근거 표시와 문장 구조", "답을 고른 문장에 근거를 표시하고 오답 선택지를 지운 이유"),
        ("지문", "독해 근거 표시와 문장 구조", "답을 고른 문장에 근거를 표시하고 오답 선택지를 지운 이유"),
    ]
    matched = [(label, evidence) for token, label, evidence in candidates if token in text]
    if matched:
        return matched[stable_number(key, "english-focus") % len(matched)]
    defaults = [
        ("독해 근거와 어휘 복습", "지문 근거 표시와 다음 수업 전 어휘 재확인 기록"),
        ("문장 구조와 서술형 표현", "문장 구조를 나눈 흔적과 서술형 답안을 고친 기록"),
        ("어휘·문법·독해의 연결", "영역별 오답을 나누고 다시 설명한 기록"),
    ]
    return defaults[stable_number(key, "english-default") % len(defaults)]


def detect_math_focus(text: str, key: str) -> tuple[str, str]:
    candidates = [
        ("계산", "계산 과정과 검산 습관", "중간 계산을 지우지 않고 실수 지점을 표시한 풀이"),
        ("서술형", "서술형 풀이의 조건과 근거", "조건에 밑줄을 긋고 식을 세운 이유를 적은 풀이"),
        ("응용", "응용 문제의 조건 해석", "낯선 조건을 단원 개념과 연결해 다시 푼 기록"),
        ("유형", "유형 변형과 재풀이", "대표 유형과 변형 문항을 나란히 놓고 차이를 설명한 기록"),
        ("개념", "개념 연결과 풀이 순서", "첫 식을 세운 이유와 다음 단계로 이어지는 개념을 설명한 기록"),
        ("오답", "오답 원인과 재풀이", "틀린 이유를 개념·조건·계산으로 나누고 다시 푼 기록"),
    ]
    matched = [(label, evidence) for token, label, evidence in candidates if token in text]
    if matched:
        return matched[stable_number(key, "math-focus") % len(matched)]
    defaults = [
        ("개념 연결과 오답 재풀이", "첫 풀이와 다시 푼 풀이를 비교한 기록"),
        ("문제 조건과 풀이 과정", "조건 표시와 식의 순서를 함께 남긴 풀이"),
        ("유형 적용과 계산 점검", "유형별 실수와 검산 결과를 나눈 기록"),
    ]
    return defaults[stable_number(key, "math-default") % len(defaults)]


def build_persona(key: str, selected_grade: str) -> str:
    learner = (
        f"{GRADE_EXPANDED[selected_grade]} 학생"
        if selected_grade
        else "영어와 수학의 공통 수업 가능 범위를 먼저 확인해야 하는 학생"
    )
    templates = [
        f"영어 지문의 근거는 찾지만 수학 풀이 조건을 끝까지 확인하지 않는 {learner}",
        f"단어와 공식은 외웠지만 새로운 문제에서 적용 순서를 잃는 {learner}",
        f"영어 복습이 밀릴 때 수학 과제도 함께 늦어지는 {learner}",
        f"답은 맞혀도 영어 문장 근거와 수학 풀이 이유를 설명하기 어려운 {learner}",
        f"시험 직전 한 과목에 시간이 몰려 다른 과목의 누적 복습이 끊기는 {learner}",
        f"과제를 끝내도 틀린 이유를 기록하지 않아 같은 실수가 이어지는 {learner}",
        f"수업 중에는 이해하지만 이틀 뒤 혼자 다시 설명하거나 풀기 어려운 {learner}",
        f"학교 진도는 따라가지만 영어와 수학의 복습 간격을 따로 정하지 못한 {learner}",
        f"공부 시간은 길지만 과목별 시작 단원과 완료 기준이 분명하지 않은 {learner}",
        f"영어 답안과 수학 풀이를 고친 뒤 재확인 날짜를 남기지 않는 {learner}",
        f"쉬운 문항의 실수가 반복되는데 개념 부족과 습관 문제를 구분하지 못한 {learner}",
        f"두 과목 중 먼저 보완할 영역을 정하지 못해 주간 계획이 자주 바뀌는 {learner}",
    ]
    return choose(key, templates, "persona")


def make_records(workbook_path: Path) -> list[Record]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    center_rows = load_csv(CENTER_CSV)
    image_rows = load_csv(IMAGE_CSV)
    cells = workbook_cells(workbook_path)
    national = load_national_sources()
    if not (len(cells) == len(center_rows) == len(image_rows) == 371):
        raise ValueError(
            f"371행이 필요합니다: workbook={len(cells)}, center={len(center_rows)}, image={len(image_rows)}"
        )

    existing_slugs = {
        normalize_slug(path.name)
        for path in (SUBJECT_ROOT / "중학생학원").iterdir()
        if path.is_dir()
    }
    records: list[Record] = []
    for index, (source_html, center, image_row) in enumerate(
        zip(cells, center_rows, image_rows, strict=True), start=1
    ):
        locality = compact(center.get("근처 수업가능 동네"))
        image_locality = compact(image_row.get("제목"))
        slug = normalize_slug(locality)
        if not locality or normalize_slug(image_locality) != slug:
            raise ValueError(f"{index}행 지역/이미지 매칭 오류: {locality!r}, {image_locality!r}")
        if slug not in existing_slugs or slug not in national:
            raise KeyError(f"{index}행 기존 페이지 매칭 실패: {locality} ({slug})")

        address = compact(center.get("센터 주소"))
        physical_region, physical_city = physical_address_parts(address)
        service_region, service_city = service_area(center, address)
        english_grades = grade_values(field(center, "가능학년", "영어"))
        math_grades = grade_values(field(center, "가능학년", "수학"))
        common_grades = [grade for grade in GRADE_ORDER if grade in english_grades and grade in math_grades]
        selected_grade = (
            common_grades[stable_number(f"{index}|{locality}", "grade") % len(common_grades)]
            if common_grades else ""
        )
        schools = school_values(
            field(center, "타깃학교", "초"),
            field(center, "타깃학교", "중"),
            field(center, "타깃학교", "고"),
        )
        source_text = compact(html.unescape(re.sub(r"<[^>]+>", " ", source_html)))
        english_focus, english_evidence = detect_english_focus(source_text, f"{index}|{locality}")
        math_focus, math_evidence = detect_math_focus(source_text, f"{index}|{locality}")

        body_stem = Path(compact(image_row.get("본문"))).stem
        body_file = f"{body_stem}.webp"
        body_mobile_file = f"{body_stem}-mobile.webp"
        body_path = ROOT / "assets" / "centers" / "common" / body_file
        body_mobile_path = ROOT / "assets" / "centers" / "common" / body_mobile_file
        source = national[slug]
        map_file = source["map_file"]
        map_path = ROOT / "assets" / "maps" / map_file
        if not body_path.exists() or not body_mobile_path.exists() or not map_path.exists():
            raise FileNotFoundError(
                f"{locality} 이미지 누락: {body_path}, {body_mobile_path}, {map_path}"
            )
        representative = source["representative"]
        if not representative.startswith("https://"):
            raise ValueError(f"{locality} 대표 이미지 URL이 없습니다")
        body_width, body_height = image_size(body_path)
        map_width, map_height = image_size(map_path)
        key = f"{index}|{locality}|{hashlib.sha256(source_text.encode()).hexdigest()}"
        records.append(
            Record(
                index=index,
                locality=locality,
                slug=slug,
                service_region=service_region,
                service_city=service_city,
                physical_region=physical_region,
                physical_city=physical_city,
                center_name=compact(center.get("센터명")),
                legal_name=compact(center.get("교육지원청명칭")),
                registration=compact(center.get("교육지원청 등록번호")),
                address=address,
                location_note=normalize_location_note(center.get("위치안내")),
                tuition_url=compact(center.get("센터 교습비")),
                english_grades=tuple(english_grades),
                math_grades=tuple(math_grades),
                common_grades=tuple(common_grades),
                schools=tuple(schools),
                source_text=source_text,
                source_html=source_html,
                source_url=source["source_url"],
                organization_id=source["organization_id"] or source["source_url"] + "#organization",
                representative=representative,
                body_file=body_file,
                body_mobile_file=body_mobile_file,
                map_file=map_file,
                body_width=body_width,
                body_height=body_height,
                map_width=map_width,
                map_height=map_height,
                english_focus=english_focus,
                english_evidence=english_evidence,
                math_focus=math_focus,
                math_evidence=math_evidence,
                selected_grade=selected_grade,
                persona=build_persona(key, selected_grade),
            )
        )
    if len({record.slug for record in records}) != 371:
        raise ValueError("371개의 고유 지역 slug가 필요합니다")
    if len({record.representative for record in records}) != 371:
        raise ValueError("기존 검증 페이지에서 371개의 고유 대표 이미지가 필요합니다")
    return records


def meta_description(record: Record) -> str:
    candidates = [
        f"{record.locality} 영수학원 상담 전 영어·수학 진단 순서, 공통 수업 가능 학년, 학교 자료, 오답·주간 계획과 센터 확인 정보를 정리했습니다.",
        f"{record.locality} 영수학원 선택에 필요한 영어·수학 학습 진단, 추천 학생, 학교 자료 활용, 주간 계획, 오답 재확인과 상담 질문을 안내합니다.",
        f"{record.locality} 영수학원을 알아보는 학부모를 위해 영어·수학의 시작점, 공통 가능 학년, 내신 자료, 복습 계획과 상담 전 체크리스트를 담았습니다.",
    ]
    value = choose(record.key, candidates, "meta")
    if not 70 <= len(value) <= 105:
        raise ValueError(f"메타 설명 길이 오류: {record.locality} {len(value)}자")
    return value


def grade_sentence(record: Record) -> str:
    if record.common_grades:
        grades = "·".join(record.common_grades)
        return (
            f"확인된 정보상 영어와 수학이 모두 가능한 학년은 {grades}입니다. "
            "실제 시간표와 학생의 현재 진도는 상담 시 다시 확인해야 합니다."
        )
    return (
        "확인된 정보만으로는 영어와 수학의 공통 수업 가능 학년을 알 수 없습니다. "
        "두 과목을 함께 알아볼 때에는 자녀 학년의 가능 여부를 상담에서 먼저 확인하세요."
    )


def school_sentence(record: Record) -> str:
    if record.schools:
        displayed = "·".join(record.schools[:8])
        return (
            f"확인된 학교 참고 정보에는 {displayed} 등이 있습니다. "
            "학교명만으로 내신 계획을 단정하지 말고 자녀가 받은 최신 범위표와 교재를 함께 대조해야 합니다."
        )
    return (
        "확인된 학교 참고 정보가 없는 지역입니다. 자녀 학교의 최근 시험 범위표, 교재, 수행평가 안내를 준비해 "
        "수업 계획에 어떻게 반영하는지 질문하세요."
    )


def content_sections(record: Record) -> tuple[str, list[dict[str, object]]]:
    answer_frames = [
        f"{record.locality} 영수학원을 비교할 때 핵심은 두 과목을 같은 진도로 묶는지가 아닙니다. 영어와 수학의 막힌 단계를 각각 설명하고, 다음 일주일에 먼저 바꿀 행동을 정하는지 확인해야 합니다.",
        f"{record.locality}에서 영수학원을 고르는 첫 기준은 수업 과목의 수보다 진단의 구체성입니다. 최근 영어 답안과 수학 풀이를 나누어 보고 과목별 복습 순서를 제시하는지 살펴보세요.",
        f"{record.locality} 영수학원 상담의 직접적인 답은 영어·수학의 현재 기록을 따로 본 뒤 한 주 계획에서 우선순위를 연결하는 곳을 찾는 것입니다. 진도보다 재현 가능한 복습 기준이 먼저입니다.",
        f"{record.locality}의 영수학원 선택에서는 영어 점수와 수학 점수를 더하는 방식보다 각 과목의 오답 원인을 구분하는 과정이 중요합니다. 진단 결과가 주간 실행표와 재확인 날짜로 이어져야 합니다.",
        f"{record.locality} 영수학원을 알아볼 때는 영어와 수학을 모두 배운다는 설명만으로 결정하기 어렵습니다. 학생 답안에서 두 과목의 출발점을 나누고 과제·복습 시간을 현실적으로 배분하는지 확인하세요.",
        f"{record.locality} 영수학원 상담에서는 어느 과목을 더 많이 공부할지보다 무엇을 먼저 고칠지 답을 얻어야 합니다. 영어 근거 표시와 수학 풀이 기록이 다음 계획에 반영되는지가 선택 기준입니다.",
        f"{record.locality}에서 영어와 수학을 함께 관리하려면 두 과목의 진단 기준과 복습 간격이 각각 보여야 합니다. 상담 뒤 학생이 실행할 첫 행동까지 설명할 수 있는지 확인하는 것이 좋습니다.",
        f"{record.locality} 영수학원 비교의 출발점은 최근 시험지와 교재입니다. 영어의 읽기·표현 과정과 수학의 조건·풀이 과정을 따로 점검한 뒤 한 주 계획으로 연결해야 합니다.",
    ]
    answer = choose(record.key, answer_frames, "answer")

    diagnosis_headings = [
        f"{record.locality} 영수학원 첫 진단에서 확인할 답",
        f"{record.locality} 영어·수학 상담의 시작점을 정하는 방법",
        f"두 과목의 현재 기록을 먼저 나누어 보는 이유",
        f"{record.locality}에서 영수학원을 비교하는 첫 기준",
        f"최근 답안으로 영어와 수학의 출발점을 찾는 순서",
        f"{record.locality} 영수학원 상담이 답해야 할 첫 질문",
    ]
    diagnosis_paragraphs = [
        f"{record.persona}에게는 진도를 서둘러 맞추기보다 최근 영어 답안과 수학 풀이를 같은 날짜 기준으로 펼쳐 보는 과정이 필요합니다. 정답 개수보다 어디에서 멈췄는지, 설명 없이 넘어간 단계가 무엇인지 먼저 구분하세요.",
        f"{record.locality}에서 확인할 학습 초점은 영어의 {with_josa(record.english_focus, '과', '와')} 수학의 {record.math_focus}입니다. 상담에서는 두 항목을 한 점수로 합치지 말고 각각의 확인 자료와 다음 행동을 요청하는 편이 정확합니다.",
    ]

    subject_headings = [
        f"{with_josa(record.english_focus, '과', '와')} {with_josa(record.math_focus, '을', '를')} 따로 점검하기",
        f"영어 답안과 수학 풀이에서 서로 다른 근거 찾기",
        f"영어·수학의 오답 원인을 과목별로 구분하는 기준",
        f"두 과목을 같은 방식으로 복습하지 않아야 하는 이유",
        f"영어의 근거와 수학의 풀이 과정을 확인하는 방법",
        f"과목별 진단 기록을 다음 학습으로 연결하기",
    ]
    subject_paragraphs = [
        f"영어에서는 {with_josa(record.english_evidence, '을', '를')} 확인하면 현재 막힘을 설명하기 쉽습니다. 수학에서는 {with_josa(record.math_evidence, '을', '를')} 함께 보면 개념 부족, 조건 해석, 계산 습관 중 무엇을 먼저 보완할지 구분할 수 있습니다.",
        f"{record.locality} 학생의 두 과목 기록은 같은 형식일 필요가 없습니다. 영어는 읽고 고친 근거를, 수학은 식과 재풀이 순서를 남기고 각 기록이 다음 수업 전 복습으로 이어지는지 확인하세요.",
    ]

    fit_headings = [
        f"{record.locality}에서 추천 학생과 가능 학년 확인하기",
        f"지역·학년·학습 상황으로 보는 영수학원 적합성",
        f"어떤 학생이 두 과목 진단부터 받아보면 좋을까요?",
        f"{record.locality} 영수학원 상담이 필요한 학습 장면",
        f"수업 가능 학년과 추천 학생을 함께 보는 이유",
        f"자녀의 현재 상황에 맞는 영어·수학 시작점",
    ]
    fit_paragraphs = [
        f"구체적으로 살펴볼 대상은 {record.persona}입니다. 같은 학년이라도 영어와 수학의 막힘은 다를 수 있으므로 학생의 최근 기록을 기준으로 필요한 설명과 연습량을 정해야 합니다.",
        grade_sentence(record),
    ]

    plan_headings = [
        f"두 과목의 주간 계획과 복습 간격을 정하는 방법",
        f"영어·수학 시간을 현실적인 한 주 계획으로 연결하기",
        f"수업 뒤 행동이 보이는 과목별 복습표 만들기",
        f"{record.locality} 학생의 오답 재확인 날짜 정하기",
        f"한 과목에 치우치지 않는 주간 실행 기준",
        f"과제·오답·재풀이를 다음 일정에 반영하는 순서",
    ]
    plan_paragraphs = [
        f"{record.persona}에게는 매일 두 과목을 같은 분량으로 배치하기보다 영어와 수학의 재확인 주기를 다르게 두는 편이 현실적입니다. 수업 당일에는 이해한 내용을 짧게 설명하고, 이틀 안에는 혼자 다시 풀거나 쓰며, 주말에는 누적 기록을 확인하세요.",
        f"{record.locality} 영수학원 상담에서는 완료한 분량만 묻지 말고 미완료 이유와 재시작 시간을 어떻게 기록하는지 확인해야 합니다. 계획이 밀리면 과목별 우선순위를 다시 정하고 다음 주에 같은 문제가 반복되는지도 살펴보세요.",
    ]

    school_headings = [
        f"{record.locality} 학교 자료를 영어·수학 계획에 반영하기",
        f"내신 준비는 학교명보다 최신 범위표에서 시작합니다",
        f"교재·시험지·수행평가 안내를 상담에 가져가는 이유",
        f"학교 일정과 두 과목 복습 계획을 맞추는 방법",
        f"{record.locality} 내신 상담에서 확인할 자료 범위",
        f"확인된 학교 정보와 실제 시험 범위를 구분하기",
    ]
    school_paragraphs = [
        school_sentence(record),
        f"{record.locality} 내신 준비에서는 영어의 지문·서술형 자료와 수학의 단원·풀이 자료를 나누어 가져가는 것이 좋습니다. 상담 답변이 학교 일정, 현재 교재, 학생 답안에 맞춰 구체화되는지 확인하세요.",
    ]

    checklist_headings = [
        f"{record.locality} 영수학원 상담 전 체크리스트",
        f"상담 시간을 구체적으로 만드는 네 가지 준비",
        f"등록을 결정하기 전에 확인할 영어·수학 질문",
        f"학부모가 상담 전에 한 장으로 정리할 내용",
        f"두 과목 수업을 비교할 때 빠뜨리지 말아야 할 항목",
        f"{record.locality} 상담에서 바로 물어볼 실행 기준",
    ]
    checklist_intro = choose(
        record.key,
        [
            f"{record.locality} 상담 전에는 다음 자료와 질문을 한 장에 정리해 두세요. 답변을 같은 기준으로 기록하면 학원마다 설명이 달라도 비교하기 쉽습니다.",
            f"영어·수학 수업을 함께 알아볼수록 준비 자료는 단순해야 합니다. {record.locality}에서는 아래 네 항목을 기준으로 진단과 실행 계획이 연결되는지 확인하세요.",
            f"상담에서 많은 질문을 던지기보다 학생의 현재 기록과 다음 행동을 확인하는 편이 유용합니다. {record.locality} 영수학원 비교에는 아래 항목이면 충분합니다.",
        ],
        "check-intro",
    )
    checklist = [
        f"최근 영어 답안과 수학 풀이에서 각각 다시 보고 싶은 문항 2~3개",
        f"학교 시험 범위와 현재 교재, 수행평가 또는 과제 일정",
        f"평일·주말에 실제로 가능한 복습 시간과 이동 시간",
        f"진단 결과, 과제 미완료, 오답 재확인을 다음 계획에 반영하는 방법",
    ]

    sections = [
        {"heading": choose(record.key, diagnosis_headings, "h1"), "paragraphs": diagnosis_paragraphs},
        {"heading": choose(record.key, subject_headings, "h2"), "paragraphs": subject_paragraphs},
        {"heading": choose(record.key, fit_headings, "h3"), "paragraphs": fit_paragraphs},
        {"heading": choose(record.key, plan_headings, "h4"), "paragraphs": plan_paragraphs},
        {"heading": choose(record.key, school_headings, "h5"), "paragraphs": school_paragraphs},
        {"heading": choose(record.key, checklist_headings, "h6"), "paragraphs": [checklist_intro], "items": checklist},
    ]
    theme = source_theme(
        record.source_html,
        record.locality,
        "영수학원",
        f"{with_josa(record.english_focus, '과', '와')} {record.math_focus}의 연결",
    )
    sections[0]["heading"] = f"{record.locality} 영수학원, {theme}"
    authored = source_paragraphs(
        record.source_html,
        useful_terms=("영어", "수학", "학습", "학생", "상담", "복습", "오답", "시험", "과제"),
        blocked_terms=("국어",),
        limit=6,
    )
    distribute_source_paragraphs(sections, authored)
    for section in sections:
        heading = str(section["heading"])
        if record.locality not in heading:
            section["heading"] = f"{record.locality} {heading}"
    if len({str(section["heading"]) for section in sections}) != 6:
        raise ValueError(f"{record.locality} H2 중복")
    return answer, sections


def build_faqs(record: Record) -> list[tuple[str, str]]:
    grade_answer = grade_sentence(record)
    schools_answer = school_sentence(record)
    questions = [
        choose(record.key, [
            f"{record.locality} 영수학원 상담에서는 영어와 수학 중 무엇부터 진단하나요?",
            f"{record.locality}에서 두 과목의 시작 순서는 어떻게 정하나요?",
            f"영어·수학 중 먼저 보완할 과목은 무엇으로 판단하나요?",
        ], "fq1"),
        choose(record.key, [
            f"{record.center_name}의 영어·수학 공통 수업 가능 학년은 어떻게 확인하나요?",
            f"{record.locality} 영수학원 수업 가능 학년은 어디에서 확인하나요?",
            f"자녀 학년이 영어와 수학 모두 가능한지는 어떻게 확인하나요?",
        ], "fq2"),
        choose(record.key, [
            f"{record.locality} 내신 상담에는 어떤 학교 자료를 준비해야 하나요?",
            f"학교별 영어·수학 시험 범위는 수업 계획에 어떻게 반영하나요?",
            f"{record.locality} 학교 정보는 영수학원 상담에서 어떻게 활용하나요?",
        ], "fq3"),
        choose(record.key, [
            f"영어와 수학의 주간 복습 시간은 같은 비율로 나누어야 하나요?",
            f"두 과목의 오답 재확인 날짜는 어떻게 다르게 정하나요?",
            f"영어·수학 과제가 함께 밀릴 때 무엇부터 조정해야 하나요?",
        ], "fq4"),
        choose(record.key, [
            f"{record.locality} 영수학원 상담 전에 꼭 물어볼 질문은 무엇인가요?",
            f"학부모가 영수학원 비교표에 남기면 좋은 항목은 무엇인가요?",
            f"첫 상담 뒤 학습 계획이 구체적인지 어떻게 판단하나요?",
        ], "fq5"),
    ]
    questions = [
        question if record.locality in question else f"{record.locality}에서 {question}"
        for question in questions
    ]
    answers = [
        f"먼저 볼 것은 점수가 더 낮은 과목이 아니라 학생이 혼자 설명하거나 다시 풀지 못하는 단계입니다. 영어는 {with_josa(record.english_evidence, '을', '를')}, 수학은 {with_josa(record.math_evidence, '을', '를')} 비교한 뒤 일주일의 첫 행동을 정하세요.",
        grade_answer,
        schools_answer,
        f"같은 비율로 고정할 필요는 없습니다. {record.persona}의 경우 과목별 오답 양, 다음 평가 일정, 혼자 복습할 수 있는 시간을 보고 영어와 수학의 분량과 재확인 날짜를 각각 정하는 편이 좋습니다.",
        f"진단에 사용한 자료, 과목별 시작 단원, 수업 뒤 복습 행동, 과제 미완료 시 조정 방법을 물어보세요. {record.locality}에서는 확인된 주소와 실제 이동 시간까지 함께 기록해야 계획을 오래 유지하기 쉽습니다.",
    ]
    answers = [f"{record.locality}에서 확인할 때, {answer}" for answer in answers]
    faqs = list(zip(questions, answers, strict=True))
    if len({question for question, _ in faqs}) != 5 or len({answer for _, answer in faqs}) != 5:
        raise ValueError(f"{record.locality} FAQ 중복")
    return faqs


def review_scenario(record: Record) -> str:
    frames = [
        f"상담 전에는 영어와 수학 중 어느 과목 점수가 더 낮은지만 보게 됐습니다. 최근 답안과 풀이 기록을 나누어 보니 {record.persona}에게 필요한 첫 행동과 다음 확인 날짜를 질문할 수 있었습니다.",
        f"두 과목을 같은 시간표로 묶으면 된다고 생각했지만, {with_josa(record.english_focus, '과', '와')} {record.math_focus}의 복습 간격이 다르다는 점을 먼저 확인했습니다. {record.locality} 상담 기준을 한 장에 적으니 비교가 구체적이었습니다.",
        f"진도와 교재만 물어보기보다 아이가 직접 설명한 영어 답안과 다시 푼 수학 풀이를 준비했습니다. 상담에서 과목별 시작점과 일주일 계획을 나누어 들으니 가정에서 확인할 내용도 분명해졌습니다.",
        f"영어 과제가 밀리면 수학 복습까지 늦어지는 이유를 단순한 의지 문제로 보지 않았습니다. 학생 기록을 기준으로 두 과목의 우선순위와 재확인 날짜를 묻는 방식이 상담 준비에 도움이 됐습니다.",
        f"학교 이름만 말하면 내신 계획이 정해질 것으로 생각했지만, 최신 범위표와 실제 답안이 더 중요했습니다. {record.locality}에서 확인할 학교 자료와 복습 질문을 미리 나누어 정리했습니다.",
        f"많은 문제를 푸는 수업보다 틀린 이유를 설명하고 다시 확인하는 흐름을 먼저 살폈습니다. 영어 근거 표시와 수학 풀이 과정이 다음 계획에 어떻게 반영되는지 묻기 쉬워졌습니다.",
    ]
    return choose(record.key, frames, "review")


def related_links(record: Record, records: list[Record], index: int) -> list[tuple[str, str, str]]:
    previous = records[(index - 1) % len(records)]
    following = records[(index + 1) % len(records)]
    return [
        ("CATEGORY", "영수학원 전체 지역", absolute_url("과목별학원", CATEGORY)),
        ("LOCAL", f"{record.locality} 지역 학원 안내", record.source_url),
        ("GRADE", f"{record.locality} 초등학생학원", absolute_url("과목별학원", "초등학생학원", record.slug)),
        ("GRADE", f"{record.locality} 중학생학원", absolute_url("과목별학원", "중학생학원", record.slug)),
        ("GRADE", f"{record.locality} 고등학생학원", absolute_url("과목별학원", "고등학생학원", record.slug)),
        ("NEARBY", previous.title, previous.page_url),
        ("NEARBY", following.title, following.page_url),
        ("GUIDE", "학습가이드", absolute_url("학습가이드")),
        ("CONTACT", "상담 준비하기", absolute_url("상담문의")),
    ]


def schema_graph(
    record: Record,
    meta: str,
    answer: str,
    sections: list[dict[str, object]],
    faqs: list[tuple[str, str]],
    related: list[tuple[str, str, str]],
) -> dict:
    page_url = record.page_url
    hub_url = absolute_url("과목별학원", CATEGORY)
    parent_url = absolute_url("과목별학원")
    body_url = BASE_URL + f"/assets/centers/common/{record.body_file}"
    map_url = BASE_URL + f"/assets/maps/{record.map_file}"
    image_id = page_url + "#primaryimage"
    article_id = page_url + "#article"
    service_id = page_url + "#service"
    about = [
        {"@type": "Thing", "name": record.title},
        {"@type": "Thing", "name": "영어·수학 학습 진단"},
        {"@type": "Thing", "name": "영어 복습과 수학 오답 재풀이"},
        {"@type": "Place", "name": record.locality},
        {"@type": "Place", "name": record.service_city},
        {"@type": "Place", "name": record.service_region},
    ]
    mentions: list[dict] = [
        {"@type": "Thing", "name": record.english_focus},
        {"@type": "Thing", "name": record.math_focus},
        {"@type": "Thing", "name": "주간 학습 계획"},
        {"@type": "Thing", "name": "오답 재확인"},
        {"@type": "Thing", "name": "내신 자료 활용"},
    ]
    mentions.extend({"@type": "EducationalOrganization", "name": school} for school in record.schools)
    offer = {
        "@type": "Offer",
        "name": f"{record.locality} 영어·수학 학습 상담",
        "url": page_url,
        "itemOffered": {
            "@type": "Service",
            "name": f"{record.locality} 영어·수학 학습 가능 범위 확인",
            "serviceType": "영어·수학 학습 상담",
        },
    }
    physical_address = {
        "@type": "PostalAddress",
        "streetAddress": record.address,
        "addressRegion": record.physical_region,
        "addressLocality": record.physical_city,
        "addressCountry": "KR",
    }
    org = {
        "@type": ["EducationalOrganization", "LocalBusiness"],
        "@id": record.organization_id,
        "name": record.center_name,
        "legalName": record.legal_name,
        "url": record.source_url,
        "address": physical_address,
        "areaServed": [
            {"@type": "Place", "name": record.locality},
            {"@type": "AdministrativeArea", "name": record.service_city},
        ],
        "knowsAbout": ["영어 학습", "수학 학습", "학습 진단", "주간 계획", "오답 재학습"],
        "makesOffer": [offer],
        "identifier": {
            "@type": "PropertyValue",
            "propertyID": "교육지원청 등록 정보",
            "value": record.registration,
        },
        "image": record.representative,
    }
    if record.tuition_url:
        org["subjectOf"] = {
            "@type": "CreativeWork",
            "name": f"{record.center_name} 교습비 안내",
            "url": record.tuition_url,
        }
    image_object = {
        "@type": "ImageObject",
        "@id": image_id,
        "url": record.representative,
        "contentUrl": record.representative,
        "caption": f"{record.title} {SITE_NAME} 대표 이미지",
        "inLanguage": "ko-KR",
    }
    breadcrumb = {
        "@type": "BreadcrumbList",
        "@id": page_url + "#breadcrumb",
        "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "홈", "item": BASE_URL + "/"},
            {"@type": "ListItem", "position": 2, "name": "과목별학원", "item": parent_url},
            {"@type": "ListItem", "position": 3, "name": CATEGORY, "item": hub_url},
            {"@type": "ListItem", "position": 4, "name": record.title, "item": page_url},
        ],
    }
    part_names = [
        "핵심 요약",
        *[str(section["heading"]) for section in sections],
        "지역·학년·추천 학생",
        "검증된 센터 정보",
        "FAQ",
        "학부모 상담 메모",
        "관련 페이지",
    ]
    has_parts = [
        {"@type": "WebPageElement", "name": name, "isPartOf": {"@id": page_url + "#webpage"}}
        for name in part_names
    ]
    webpage = {
        "@type": "WebPage",
        "@id": page_url + "#webpage",
        "url": page_url,
        "name": record.title,
        "description": meta,
        "inLanguage": "ko-KR",
        "isPartOf": {"@id": BASE_URL + "/#website"},
        "publisher": {"@id": record.organization_id},
        "primaryImageOfPage": {"@id": image_id},
        "breadcrumb": {"@id": page_url + "#breadcrumb"},
        "mainEntity": {"@id": service_id},
        "about": about,
        "mentions": mentions,
        "hasPart": has_parts,
        "dateModified": TODAY,
    }
    article = {
        "@type": "Article",
        "@id": article_id,
        "headline": record.title,
        "description": meta,
        "abstract": answer,
        "inLanguage": "ko-KR",
        "mainEntityOfPage": {"@id": page_url + "#webpage"},
        "author": {"@id": record.organization_id},
        "publisher": {"@id": record.organization_id},
        "datePublished": PUBLISHED_DATE,
        "dateModified": TODAY,
        "image": [record.representative, body_url, map_url],
        "articleSection": [
            CATEGORY, record.service_region, record.service_city, record.locality,
            *[str(section["heading"]) for section in sections],
        ],
        "about": about,
        "mentions": mentions,
        "hasPart": has_parts,
    }
    service = {
        "@type": "Service",
        "@id": service_id,
        "name": f"{record.title} 학습 상담 안내",
        "serviceType": "영어·수학 학습 진단 및 계획 상담",
        "description": answer,
        "provider": {"@id": record.organization_id},
        "areaServed": {"@type": "Place", "name": record.locality},
        "audience": {
            "@type": "EducationalAudience",
            "educationalRole": "student",
            "audienceType": "영어·수학 학습 상담 대상",
        },
        "about": about,
        "mentions": mentions,
        "makesOffer": [offer],
    }
    faq_node = {
        "@type": "FAQPage",
        "@id": page_url + "#faq",
        "mainEntity": [
            {
                "@type": "Question",
                "name": question,
                "acceptedAnswer": {"@type": "Answer", "text": answer_text},
            }
            for question, answer_text in faqs
        ],
    }
    related_node = {
        "@type": "ItemList",
        "@id": page_url + "#related",
        "name": f"{record.locality} 관련 학원·학습 안내",
        "numberOfItems": len(related),
        "itemListElement": [
            {"@type": "ListItem", "position": position, "name": label, "url": url}
            for position, (_, label, url) in enumerate(related, 1)
        ],
    }
    graph: list[dict] = [
        {
            "@type": "WebSite",
            "@id": BASE_URL + "/#website",
            "url": BASE_URL + "/",
            "name": SITE_NAME,
            "inLanguage": "ko-KR",
            "publisher": {"@id": BASE_URL + "/#organization"},
        },
        org,
        webpage,
        image_object,
        breadcrumb,
        article,
        service,
        faq_node,
        related_node,
    ]
    if record.schools:
        graph.append(
            {
                "@type": "ItemList",
                "@id": page_url + "#schools",
                "name": f"{record.locality} 학교 참고 정보",
                "numberOfItems": len(record.schools),
                "itemListElement": [
                    {"@type": "ListItem", "position": position, "name": school}
                    for position, school in enumerate(record.schools, 1)
                ],
            }
        )
    return {"@context": "https://schema.org", "@graph": graph}


def root_nav(active: str) -> str:
    links = [
        ("홈", "/"), ("진단상담", "/진단상담/"),
        ("학습가이드", "/학습가이드/"), ("전국학원", "/전국학원/"),
        ("과목별학원", "/과목별학원/"), ("상담문의", "/상담문의/"),
    ]
    rendered = "".join(
        f'<a{" class=\"active\" aria-current=\"page\"" if label == active else ""} href="{href}">{label}</a>'
        for label, href in links
    )
    return f'''<header class="site-header">
    <nav class="nav" aria-label="주요 메뉴">
      <a class="brand" href="/"><span class="brand-mark">L</span><span>{SITE_NAME}</span></a>
      <div class="nav-links">{rendered}</div>
      <a class="nav-cta" href="/상담문의/">상담 신청</a>
    </nav>
  </header>'''


def footer() -> str:
    return f'''<footer class="site-footer"><div class="wrap footer-inner">
    <strong>{SITE_NAME}</strong>
    <div class="footer-links"><a href="/학습가이드/">학습가이드</a><a href="/전국학원/">전국학원</a><a href="/과목별학원/">과목별학원</a></div>
    <div class="footer-contact"><span>상담 전화</span><a href="tel:{PHONE}">{PHONE}</a></div>
  </div></footer>
  <div class="floating-actions" aria-label="빠른 상담 메뉴">
    <a href="tel:{PHONE}" class="fab-call"><span class="fab-icon">&#128222;</span><span class="fab-text">전화문의</span></a>
    <a href="{SMS_URL}" target="_blank" rel="noopener" class="fab-sms"><span class="fab-icon">&#128172;</span><span class="fab-text">문자문의</span></a>
    <a href="{FORM_URL}" target="_blank" rel="noopener" class="fab-consult pulse-effect"><span class="fab-icon">&#128221;</span><span class="fab-text">상담신청</span></a>
  </div>'''


def render_sections(sections: list[dict[str, object]]) -> str:
    rendered: list[str] = []
    for section in sections:
        paragraphs = "".join(f"<p>{esc(value)}</p>" for value in section["paragraphs"])
        items = section.get("items") or []
        item_html = ""
        if items:
            item_html = '<ul class="subject-copy-list">' + "".join(
                f"<li>{esc(value)}</li>" for value in items
            ) + "</ul>"
        rendered.append(
            f'<section class="subject-copy-section"><h2>{esc(section["heading"])}</h2>{paragraphs}{item_html}</section>'
        )
    return "\n".join(rendered)


def render_facts(record: Record) -> str:
    english = "·".join(record.english_grades) if record.english_grades else "상담 시 확인"
    math = "·".join(record.math_grades) if record.math_grades else "상담 시 확인"
    common = "·".join(record.common_grades) if record.common_grades else "상담 확인 필요"
    school_tags = (
        "".join(f"<span>{esc(school)}</span>" for school in record.schools)
        if record.schools else "<span>학교 자료를 상담 시 확인</span>"
    )
    tuition = (
        f'<dt>교습비 자료</dt><dd><a href="{esc(record.tuition_url)}" target="_blank" rel="noopener noreferrer">센터별 안내 확인 <span aria-hidden="true">↗</span></a></dd>'
        if record.tuition_url else "<dt>교습비 자료</dt><dd>상담 시 확인</dd>"
    )
    location = (
        f'<dt>위치 참고</dt><dd>{esc(record.location_note)}</dd>'
        if record.location_note else ""
    )
    return f'''<section class="center-profile-overview yeongsu-fit-section" aria-labelledby="fit-facts-title">
      <div class="wrap center-profile-overview-grid">
        <div><p class="subject-kicker">LOCAL · GRADE · STUDENT</p><h2 id="fit-facts-title">{esc(record.locality)} 지역·학년·추천 학생</h2><p>{esc(record.persona)}을 구체적인 상담 대상으로 삼았습니다. 학년 범위와 실제 반 편성은 아래 확인 정보와 현재 상담 내용을 함께 보세요.</p></div>
        <dl class="center-profile-facts"><dt>센터 기준</dt><dd>{esc(record.center_name)}</dd><dt>확인된 주소</dt><dd>{esc(record.address)}</dd>{location}<dt>영어 가능 학년</dt><dd>{esc(english)}</dd><dt>수학 가능 학년</dt><dd>{esc(math)}</dd><dt>영어·수학 공통 가능 학년</dt><dd>{esc(common)}</dd><dt>등록 학원명</dt><dd>{esc(record.legal_name)}</dd><dt>등록 정보</dt><dd>{esc(record.registration)}</dd>{tuition}</dl>
      </div>
    </section>
    <section class="center-profile-school"><div class="wrap center-profile-school-grid"><div><p class="subject-kicker">SCHOOL REFERENCE</p><h2>{esc(record.locality)} 학교 참고 정보</h2><p>확인된 명칭만 표시했습니다. 실제 내신 범위는 학생이 받은 최신 교재와 학교 자료로 다시 확인하세요.</p></div><div class="center-profile-school-list">{school_tags}</div></div></section>'''


def render_local_page(record: Record, records: list[Record], index: int) -> str:
    meta = meta_description(record)
    answer, sections = content_sections(record)
    faqs = build_faqs(record)
    related = related_links(record, records, index)
    graph = schema_graph(record, meta, answer, sections, faqs, related)
    body_src = f"../../../assets/centers/common/{record.body_file}"
    body_mobile_src = f"../../../assets/centers/common/{record.body_mobile_file}"
    body_mobile_avif = f"../../../assets/generated/yeongsu-{Path(record.body_mobile_file).stem}.avif"
    map_src = f"../../../assets/maps/{record.map_file}"
    faq_html = "\n".join(
        f'<details class="subject-faq-item"{" open" if position == 1 else ""}><summary><span>Q</span>{esc(question)}</summary><div class="subject-faq-answer"><span>A</span><p>{esc(answer_text)}</p></div></details>'
        for position, (question, answer_text) in enumerate(faqs, 1)
    )
    related_html = "\n".join(
        f'<a href="{esc(url)}"><span>{esc(kind)}</span><strong>{esc(label)}</strong><i aria-hidden="true">→</i></a>'
        for kind, label, url in related
    )
    review = review_scenario(record)
    schema = json.dumps(graph, ensure_ascii=False, separators=(",", ":"))
    return f'''<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(record.title)} | {SITE_NAME}</title>
  <meta name="description" content="{esc(meta)}">
  <meta name="robots" content="index, follow, max-image-preview:large">
  <link rel="canonical" href="{record.page_url}">
  <meta property="og:locale" content="ko_KR">
  <meta property="og:site_name" content="{SITE_NAME}">
  <meta property="og:type" content="article">
  <meta property="og:title" content="{esc(record.title)} | {SITE_NAME}">
  <meta property="og:description" content="{esc(meta)}">
  <meta property="og:url" content="{record.page_url}">
  <meta property="og:image" content="{esc(record.representative)}">
  <meta property="og:image:alt" content="{esc(record.title)} {SITE_NAME} 대표 이미지">
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:title" content="{esc(record.title)} | {SITE_NAME}">
  <meta name="twitter:description" content="{esc(meta)}">
  <meta name="twitter:image" content="{esc(record.representative)}">
  <link rel="alternate" type="application/rss+xml" title="{SITE_NAME} RSS" href="{BASE_URL}/rss.xml">
  <link rel="icon" type="image/png" href="../../../assets/favicon.png">
  <link rel="stylesheet" href="../../../assets/subject.css">
  <script type="application/ld+json">{schema}</script>
</head>
<body class="subject-academy-page yeongsu-subject-page">
  <a class="skip-link" href="#main">본문 바로가기</a>
  {root_nav("과목별학원")}
  <main id="main">
    <section class="subject-local-hero"><div class="wrap">
      <nav class="subject-breadcrumb" aria-label="현재 위치"><a href="/">홈</a><span>›</span><a href="/과목별학원/">과목별학원</a><span>›</span><a href="/과목별학원/영수학원/">영수학원</a><span>›</span><strong>{esc(record.title)}</strong></nav>
      <p class="subject-kicker">ENGLISH &amp; MATH COACHING · {esc(record.service_region)} {esc(record.service_city)}</p>
      <h1>{esc(record.title)}</h1>
      <p class="subject-hero-answer">{esc(meta)}</p>
      <div class="subject-hero-tags"><span>{esc(record.service_region)}</span><span>{esc(record.service_city)}</span><span>영어·수학</span><span>진단·계획·오답 재확인</span></div>
    </div></section>
    <section class="subject-quick-answer" aria-label="{esc(record.title)} 핵심 답변"><div class="wrap subject-narrow"><div class="subject-answer-box"><span>핵심 답변</span><p>{esc(answer)}</p></div></div></section>
    <section class="subject-media-section"><div class="wrap">
      <img class="subject-hidden-representative" data-role="representative-image" src="{esc(record.representative)}" alt="{esc(record.title)} {SITE_NAME} 대표 이미지" style="display:none;">
      <figure class="subject-body-card"><div class="subject-media-label"><span>01</span><strong>{esc(record.locality)} 영어·수학 학습 안내</strong></div><picture><source media="(max-width:720px)" type="image/avif" srcset="{body_mobile_avif}"><source media="(max-width:720px)" type="image/webp" srcset="{body_mobile_src}"><img src="{body_src}" alt="{esc(record.title)} 영어·수학 학습 안내 이미지" width="{record.body_width}" height="{record.body_height}" fetchpriority="high" decoding="async"></picture></figure>
      <figure class="subject-map-card"><div class="subject-media-label"><span>02</span><strong>{esc(record.locality)} 위치 안내</strong></div><img src="{map_src}" alt="{esc(record.title)} 센터 위치 지도" width="{record.map_width}" height="{record.map_height}" loading="lazy" decoding="async"><figcaption>확인된 센터 주소는 {esc(record.address)}입니다. 방문 가능 시간은 상담 시 확인하세요.</figcaption></figure>
    </div></section>
    <article class="subject-manuscript wrap" aria-labelledby="manuscript-title">
      <header class="subject-copy-head"><p>ANSWER-FIRST ENGLISH &amp; MATH GUIDE</p><h2 id="manuscript-title">{esc(record.title)} 선택 전 확인할 학습 기준</h2></header>
      <div class="subject-copy-flow">{render_sections(sections)}</div>
    </article>
    {render_facts(record)}
    <section class="subject-faq-section"><div class="wrap subject-narrow"><div class="subject-section-head"><p>QUESTIONS &amp; ANSWERS</p><h2>{esc(record.title)} 자주 묻는 질문</h2><span>화면의 질문과 답변은 구조화 데이터와 동일합니다.</span></div><div class="subject-faq-list">{faq_html}</div></div></section>
    <section class="subject-review-section"><div class="wrap subject-narrow"><div class="subject-review-card"><p class="subject-review-label">PARENT VOICE GUIDE</p><h2>{esc(record.title)} 학부모 상담 메모</h2><blockquote>{esc(review)}</blockquote><p class="subject-review-note">특정 학부모의 이용 경험이나 성적 사례가 아니라, 대표적인 상담 준비 상황을 재구성한 예시입니다.</p></div></div></section>
    <section class="subject-related-section"><div class="wrap subject-narrow"><div class="subject-section-head"><p>RELATED PAGES</p><h2>{esc(record.locality)} 학원 정보 이어보기</h2><span>같은 동네의 학년별 안내와 인접한 영수학원 페이지를 함께 확인할 수 있습니다.</span></div><div class="subject-related-grid">{related_html}</div></div></section>
    <section class="consult-strip"><div class="wrap consult-strip-inner"><div><p class="eyebrow">상담 전 체크</p><h2>{esc(record.locality)} 최근 영어 답안과 수학 풀이 준비</h2><p>두 과목의 막힌 지점을 따로 확인해야 첫 주의 우선순위를 구체적으로 정할 수 있습니다.</p></div><a class="btn btn-primary" href="/상담문의/">상담 준비하기</a></div></section>
  </main>
  {footer()}
</body>
</html>'''


def hub_faqs() -> list[tuple[str, str]]:
    return [
        (
            "영수학원은 영어와 수학을 같은 방식으로 공부하는 곳인가요?",
            "두 과목을 함께 관리하더라도 진단과 복습 방식은 달라야 합니다. 영어는 어휘·문장·독해 근거를, 수학은 개념·조건·풀이 과정을 따로 확인한 뒤 한 주 계획에서 우선순위를 연결하는 방식이 좋습니다.",
        ),
        (
            "지역별 영수학원 페이지의 가능 학년은 어떤 자료를 기준으로 하나요?",
            "공통 센터자료에서 영어와 수학에 각각 표시된 가능 학년의 교집합만 공통 학년으로 안내합니다. 한 과목의 정보가 비어 있으면 학년을 임의로 만들지 않고 상담 확인 대상으로 표시합니다.",
        ),
        (
            "상담 전에 어떤 자료를 준비하면 지역 페이지를 활용하기 좋나요?",
            "최근 영어 답안과 수학 풀이, 현재 교재, 학교 시험 범위, 실제 복습 가능 시간을 준비하세요. 지역 페이지의 체크리스트와 센터 확인 정보를 함께 보면 상담 질문을 같은 기준으로 비교할 수 있습니다.",
        ),
    ]


def render_hub(records: list[Record]) -> str:
    hub_url = absolute_url("과목별학원", CATEGORY)
    parent_url = absolute_url("과목별학원")
    description = "전국 371개 동네별 영수학원 안내입니다. 영어·수학 진단, 공통 가능 학년, 학교 자료, 주간 계획, 오답 재확인과 상담 기준을 지역별로 확인하세요."
    grouped: dict[str, dict[str, list[Record]]] = defaultdict(lambda: defaultdict(list))
    for record in records:
        grouped[record.service_region][record.service_city].append(record)
    region_html: list[str] = []
    region_keys = sorted(
        grouped,
        key=lambda region: (REGION_ORDER.index(region) if region in REGION_ORDER else 99, region),
    )
    for region in region_keys:
        city_html: list[str] = []
        for city in sorted(grouped[region]):
            local_records = sorted(grouped[region][city], key=lambda item: item.locality)
            buttons = "".join(
                f'<a class="subject-local-button" data-local-name="{esc(item.locality)}" data-search="{esc(" ".join([item.locality, item.service_city, item.center_name]))}" href="/과목별학원/영수학원/{quote(item.slug)}/"><strong>{esc(item.locality)}</strong><span>영수학원</span></a>'
                for item in local_records
            )
            city_html.append(
                f'<section class="subject-city-group" data-city-group><h3>{esc(city)} <small>{len(local_records)}</small></h3><div class="subject-local-grid">{buttons}</div></section>'
            )
        count = sum(len(values) for values in grouped[region].values())
        opened = " open" if region == "서울" else ""
        region_html.append(
            f'<details class="subject-region-group" data-region-group{opened}><summary><span>{esc(region)}</span><strong>{count}개 지역</strong></summary><div class="subject-region-content">{"".join(city_html)}</div></details>'
        )
    faqs = hub_faqs()
    faq_html = "".join(
        f'<details class="subject-faq-item"{" open" if position == 1 else ""}><summary><span>Q</span>{esc(question)}</summary><div class="subject-faq-answer"><span>A</span><p>{esc(answer)}</p></div></details>'
        for position, (question, answer) in enumerate(faqs, 1)
    )
    graph = {
        "@context": "https://schema.org",
        "@graph": [
            {
                "@type": "CollectionPage", "@id": hub_url + "#webpage", "url": hub_url,
                "name": "영수학원 지역 안내", "description": description, "inLanguage": "ko-KR",
                "about": ["영수학원", "영어·수학 학습 진단", "주간 계획", "오답 재확인"],
                "breadcrumb": {"@id": hub_url + "#breadcrumb"},
                "hasPart": [{"@id": hub_url + "#local-list"}, {"@id": hub_url + "#faq"}],
            },
            {
                "@type": "BreadcrumbList", "@id": hub_url + "#breadcrumb",
                "itemListElement": [
                    {"@type": "ListItem", "position": 1, "name": "홈", "item": BASE_URL + "/"},
                    {"@type": "ListItem", "position": 2, "name": "과목별학원", "item": parent_url},
                    {"@type": "ListItem", "position": 3, "name": CATEGORY, "item": hub_url},
                ],
            },
            {
                "@type": "ItemList", "@id": hub_url + "#local-list",
                "name": "전국 영수학원 지역 페이지", "numberOfItems": len(records),
                "itemListElement": [
                    {"@type": "ListItem", "position": position, "name": record.title, "url": record.page_url}
                    for position, record in enumerate(records, 1)
                ],
            },
            {
                "@type": "FAQPage", "@id": hub_url + "#faq",
                "mainEntity": [
                    {"@type": "Question", "name": question, "acceptedAnswer": {"@type": "Answer", "text": answer}}
                    for question, answer in faqs
                ],
            },
        ],
    }
    schema = json.dumps(graph, ensure_ascii=False, separators=(",", ":"))
    return f'''<!doctype html>
<html lang="ko"><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
  <title>영수학원 지역 안내 | {SITE_NAME}</title>
  <meta name="description" content="{esc(description)}"><meta name="robots" content="index, follow">
  <link rel="canonical" href="{hub_url}"><meta property="og:locale" content="ko_KR"><meta property="og:site_name" content="{SITE_NAME}"><meta property="og:type" content="website"><meta property="og:title" content="영수학원 지역 안내 | {SITE_NAME}"><meta property="og:description" content="{esc(description)}"><meta property="og:url" content="{hub_url}">
  <link rel="alternate" type="application/rss+xml" title="{SITE_NAME} RSS" href="{BASE_URL}/rss.xml">
  <link rel="icon" type="image/png" href="../../assets/favicon.png"><link rel="stylesheet" href="../../assets/subject.css"><script type="application/ld+json">{schema}</script>
</head><body class="subject-hub-page yeongsu-hub-page"><a class="skip-link" href="#main">본문 바로가기</a>{root_nav("과목별학원")}<main id="main">
  <section class="subject-hub-hero"><div class="wrap"><nav class="subject-breadcrumb" aria-label="현재 위치"><a href="/">홈</a><span>›</span><a href="/과목별학원/">과목별학원</a><span>›</span><strong>영수학원</strong></nav><p class="subject-kicker">ENGLISH &amp; MATH ACADEMY DIRECTORY</p><h1>동네별 영수학원 안내</h1><p>{esc(description)}</p><div class="subject-hub-stats"><span><strong>371</strong>지역 상세 안내</span><span><strong>검증</strong>학년·학교·주소</span><span><strong>5</strong>페이지별 FAQ</span></div></div></section>
  <section class="subject-directory-section"><div class="wrap"><div class="subject-directory-head"><div><p>LOCAL DIRECTORY</p><h2>지역명으로 영수학원 찾기</h2></div><label class="subject-search"><span class="sr-only">지역명 검색</span><input id="subject-local-search" type="search" placeholder="예: 명일동, 불당동" autocomplete="off"><button id="subject-search-reset" type="button" hidden>초기화</button></label></div><p id="subject-search-status" class="subject-search-status" aria-live="polite"></p><div id="subject-region-list">{"".join(region_html)}</div></div></section>
  <section class="subject-hub-guide"><div class="wrap"><div class="subject-section-head"><p>SELECTION GUIDE</p><h2>영어와 수학은 진단과 복습 기준을 따로 확인하세요</h2></div><div class="subject-guide-grid"><article><span>01</span><h3>과목별 현재 진단</h3><p>영어 답안과 수학 풀이에서 각각 멈춘 단계를 찾아 시작점을 나눕니다.</p></article><article><span>02</span><h3>학교 자료와 학년 확인</h3><p>공통 가능 학년과 최신 시험 범위·교재를 함께 대조합니다.</p></article><article><span>03</span><h3>주간 계획과 재확인</h3><p>두 과목의 복습 간격을 다르게 두고 다음 확인 날짜를 정합니다.</p></article></div></div></section>
  <section class="subject-faq-section"><div class="wrap subject-narrow"><div class="subject-section-head"><p>FAQ</p><h2>영수학원 지역 안내 자주 묻는 질문</h2><span>지역 상세 페이지를 보기 전에 확인할 기준입니다.</span></div><div class="subject-faq-list">{faq_html}</div></div></section>
</main>{footer()}<script>
(() => {{
  const input = document.getElementById('subject-local-search');
  const reset = document.getElementById('subject-search-reset');
  const status = document.getElementById('subject-search-status');
  const cards = [...document.querySelectorAll('[data-local-name]')];
  const cities = [...document.querySelectorAll('[data-city-group]')];
  const regions = [...document.querySelectorAll('[data-region-group]')];
  const normalize = value => value.toLowerCase().replace(/\\s+/g, '');
  const update = () => {{
    const query = normalize(input.value.trim());
    let count = 0;
    cards.forEach(card => {{
      const matched = !query || normalize(card.dataset.search || card.dataset.localName).includes(query);
      card.hidden = !matched;
      if (matched) count += 1;
    }});
    cities.forEach(city => {{ city.hidden = ![...city.querySelectorAll('[data-local-name]')].some(card => !card.hidden); }});
    regions.forEach(region => {{
      region.hidden = ![...region.querySelectorAll('[data-local-name]')].some(card => !card.hidden);
      if (query && !region.hidden) region.open = true;
    }});
    reset.hidden = !query;
    status.textContent = query ? `${{count}}개 지역을 찾았습니다.` : '';
  }};
  input.addEventListener('input', update);
  reset.addEventListener('click', () => {{ input.value = ''; update(); input.focus(); }});
}})();
</script></body></html>'''


def preflight(records: list[Record]) -> None:
    for asset in (
        ROOT / "assets" / "generated" / "yeongsu-seoul-mobile.avif",
        ROOT / "assets" / "generated" / "yeongsu-local-mobile.avif",
    ):
        if not asset.is_file() or asset.stat().st_size <= 0:
            raise FileNotFoundError(f"모바일 본문 이미지가 없습니다: {asset}")
    if len(records) != 371:
        raise ValueError(f"상세 371개가 필요합니다: {len(records)}")
    titles = [record.title for record in records]
    metas = [meta_description(record) for record in records]
    urls = [record.page_url for record in records]
    if len(set(titles)) != 371 or len(set(metas)) != 371 or len(set(urls)) != 371:
        raise ValueError("title/meta/canonical 고유성 검사 실패")
    forbidden = re.compile(
        r"LOCAL ACADEMY GUIDE|핵심 키워드|(?<![가-힣])원고(?![가-힣])|이 페이지|수업 진행방식|"
        r"따라가며도|바뀌도|학습예습|영수국|실시간 수업|온라인 수업|"
        r"입시합격|합격전략|후기 기반|실제 후기|성적이 향상|점수가 올랐|"
        r"풀이을|적용와|적용를|표현와|표현를|기록를|기준를|기준는|"
        r"과정를|습관를|내용를|계획를|학생 학생|상담 상담|확인 확인",
        re.I,
    )
    for index, record in enumerate(records):
        answer, sections = content_sections(record)
        faqs = build_faqs(record)
        output = " ".join(
            [answer, record.persona, review_scenario(record)]
            + [str(section["heading"]) for section in sections]
            + [str(value) for section in sections for value in section["paragraphs"]]
            + [value for question_answer in faqs for value in question_answer]
        )
        match = forbidden.search(output)
        if match:
            raise ValueError(f"{record.locality} 공개 문구 금칙어: {match.group(0)}")
        graph = schema_graph(
            record,
            meta_description(record),
            answer,
            sections,
            faqs,
            related_links(record, records, index),
        )
        types = {
            schema_type
            for node in graph["@graph"]
            for schema_type in (
                node.get("@type", [])
                if isinstance(node.get("@type"), list)
                else [node.get("@type")]
            )
        }
        required = {
            "EducationalOrganization", "LocalBusiness", "WebPage", "ImageObject",
            "Article", "Service", "FAQPage", "BreadcrumbList", "ItemList",
        }
        if not required.issubset(types):
            raise ValueError(f"{record.locality} 스키마 누락: {sorted(required - types)}")


def write_site(records: list[Record]) -> None:
    resolved_target = TARGET_ROOT.resolve()
    resolved_subject = SUBJECT_ROOT.resolve()
    if resolved_target.parent != resolved_subject:
        raise RuntimeError(f"안전하지 않은 대상 경로: {resolved_target}")
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    expected = {record.slug for record in records}
    for child in TARGET_ROOT.iterdir():
        if child.is_dir() and child.name not in expected:
            if child.resolve().parent != resolved_target:
                raise RuntimeError(f"안전하지 않은 잔여 경로: {child}")
            shutil.rmtree(child)
    for index, record in enumerate(records):
        output = TARGET_ROOT / record.slug / "index.html"
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(
            render_local_page(record, records, index).rstrip() + "\n",
            encoding="utf-8",
            newline="\n",
        )
    (TARGET_ROOT / "index.html").write_text(
        render_hub(records).rstrip() + "\n", encoding="utf-8", newline="\n"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="학습코칭.kr 영수학원 371개 지역 페이지 생성")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--check-only", action="store_true", help="생성 전 검사만 실행")
    args = parser.parse_args()
    records = make_records(args.workbook)
    preflight(records)
    if not args.check_only:
        write_site(records)
    print(
        json.dumps(
            {
                "records": len(records),
                "unique_centers": len({record.center_name for record in records}),
                "missing_common_grades": sum(not record.common_grades for record in records),
                "missing_tuition_links": sum(not record.tuition_url for record in records),
                "written": not args.check_only,
                "target": str(TARGET_ROOT),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
